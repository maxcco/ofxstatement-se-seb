from typing import Iterable
import itertools
import re

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import  BankAccount, Statement, StatementLine, generate_transaction_id

from openpyxl import load_workbook
from openpyxl.cell import Cell

class SwedenSebPlugin(Plugin):
    def get_parser(self, filename: str) -> "SebParser":
        return SebParser(filename)


class SebParser(StatementParser[str]):

    date_format = "%Y-%m-%d"
    bank_id = 'SE31500000000'
    currency_id = 'SEK'

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename
        self.wb = load_workbook(filename=filename, read_only=True).active

    def parse(self) -> Statement:

        #extract account_id from the string in A5 using regexp
        self.account_id = re.findall(r"\((\d+)\)", wb['A5'])
        self.bank_account = BankAccount(
            bank_id=self.bank_id, self.account_id=statement.account_id
        )

        statement = super().parse()
        statement.bank_id = self.bank_id
        statement.currency = self.currency_id
        statement.account_id = self.account_id

        rows_iter = itertools.islice(self.wb.iter_rows(), 8, None)
        first_row = take(6, next(rows_iter))
        last_row = take(6, deque(rows_iter, maxlen=1).pop())

        statement.start_date= self.parse_datetime(last_row[0])
        statement.start_balance = last_row[5]
        statement.end_date = self.parse_datetime(first_row[0])
        statement.end_balance = first_row[5]

        return statement


    def split_records(self) -> Iterable[Iterable[Cell]]:

        # Skip first 8 headers rows
        for row in itertools.islice(self.wb.iter_rows(), 8, None):
            yield [c.value for c in row]

    def parse_record(self, line: Iterable[Cell]) -> StatementLine:
        col = take(6, row)

        stmt_line = StatementLine()
        stmt_line.date = self.parse_datetime(col[0])
        stmt.date_user = self.parse_datetime(col[1])
        stmt_line.refnum = col[2]
        stmt_line.memo = col[3]
        stmt_line.amount = col[4]
        stmt_line.bank_account_to = self.bank_account
        stmt_line.id = generate_transaction_id(stmt_line)

        return stmt_line
